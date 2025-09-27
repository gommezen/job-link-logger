import os
import re
import json
import base64
import datetime as dt
from typing import List, Tuple, Optional
import urllib.parse as urlparse

from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import html2text
# from urllib.parse import urlparse as _urlp  # <- (4) Optional "Source" column helper

# =========================
# Config
# =========================
SCOPES = ["https://www.googleapis.com/auth/gmail.readonly"]
EXCEL_PATH = "job_links.xlsx"
STATE_PATH = "processed.json"   # keeps processed Gmail message IDs

# (1) COMBINED GMAIL QUERY: your label OR any email that contains job links (last 60d)
LABEL_NAME = "Jobs/LinkedIn"
SEARCH_TERMS = [
    "linkedin.com/jobs",
    "lnkd.in/",
    "jobindex.dk/vis-job",
]
_terms = " OR ".join(f'"{t}"' for t in SEARCH_TERMS)
GMAIL_QUERY = f'(label:"{LABEL_NAME}" OR ({_terms})) newer_than:60d'

# (2) UNIFIED URL MATCHER: LinkedIn jobs, lnkd.in shortlinks, Jobindex postings
JOB_URL_REGEX = re.compile(
    r'(?:'
    r'https?://(?:www\.)?linkedin\.com[^\s"\'<)]+'
    r'|https?://lnkd\.in/[A-Za-z0-9_-]+'
    r'|https?://(?:www\.)?jobindex\.dk/vis-job/[^\s"\'<)]+'
    r')',
    re.IGNORECASE
)

# =========================
# Gmail auth & client
# =========================
def get_gmail_service():
    creds = None
    if os.path.exists("token.json"):
        creds = Credentials.from_authorized_user_file("token.json", SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            from google.auth.transport.requests import Request
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file("credentials.json", SCOPES)
            creds = flow.run_local_server(port=0)
        with open("token.json", "w") as token:
            token.write(creds.to_json())
    return build("gmail", "v1", credentials=creds)

# =========================
# Utility: Excel helpers
# =========================
# If you later enable the "Source" column, change "HEADERS" and the widths accordingly.
HEADERS = ["Date", "From", "Subject", "LinkedIn URL", "Gmail Permalink", "Status", "Notes"]
# HEADERS = ["Date", "From", "Subject", "Job URL", "Gmail Permalink", "Source", "Status", "Notes"]  # <- (4) Optional

def ensure_excel(path: str) -> None:
    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Links"
        ws.append(HEADERS)
        widths = [22, 38, 50, 85, 60, 14, 40]
        # widths = [22, 38, 50, 85, 60, 18, 14, 40]  # <- (4) Optional: if you add "Source"
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # Data sheet with status options
        data = wb.create_sheet("Data")
        statuses = ["", "To Review", "Applied", "Interview", "Offer", "Rejected", "On Hold"]
        data["A1"] = "Statuses"
        for i, s in enumerate(statuses, start=2):
            data[f"A{i}"] = s

        # Data validation (dropdown) on Status column (F)
        # If you enable "Source", Status becomes column G → change to "G2:G10000"
        dv = DataValidation(type="list", formula1="=Data!$A$2:$A$8", allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"F2:F10000")
        # dv.add(f"G2:G10000")  # <- (4) Optional if Source is enabled

        wb.save(path)

def read_existing_urls(path: str) -> set:
    wb = load_workbook(path)
    ws = wb["Links"]
    urls = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[3]:
            urls.add(str(row[3]).strip())
    return urls

def append_rows(path: str, rows: List[List[str]]) -> None:
    wb = load_workbook(path)
    ws = wb["Links"]
    for r in rows:
        ws.append(r)
    wb.save(path)

# =========================
# Utility: state (processed message IDs)
# =========================
def load_state(path: str) -> dict:
    if os.path.exists(path):
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    return {"processed_ids": []}

def save_state(path: str, state: dict) -> None:
    with open(path, "w", encoding="utf-8") as f:
        json.dump(state, f, ensure_ascii=False, indent=2)

# =========================
# Gmail message parsing
# =========================
def get_message(service, msg_id: str) -> dict:
    return service.users().messages().get(userId="me", id=msg_id, format="full").execute()

def decode_b64(data: str) -> str:
    return base64.urlsafe_b64decode(data.encode("utf-8")).decode("utf-8", errors="ignore")

def extract_headers(payload_headers: List[dict], name: str) -> Optional[str]:
    for h in payload_headers:
        if h.get("name", "").lower() == name.lower():
            return h.get("value")
    return None

def flattened_bodies(payload: dict) -> Tuple[str, str]:
    plain_parts, html_parts = [], []
    def walk(part):
        mime = part.get("mimeType", "")
        body = part.get("body", {})
        data = body.get("data")
        if data:
            text = decode_b64(data)
            if mime == "text/plain":
                plain_parts.append(text)
            elif mime == "text/html":
                html_parts.append(text)
        if "parts" in part:
            for p in part["parts"]:
                walk(p)
    walk(payload)
    return ("\n".join(plain_parts), "\n".join(html_parts))

# (3) NEW EXTRACTOR: handle LinkedIn, lnkd.in, and Jobindex; unwrap redirects; normalize
def extract_job_urls(subject: str, plain: str, html: str) -> List[str]:
    candidates = [subject or "", plain or ""]
    if html:
        candidates.append(html)
        candidates.append(html2text.html2text(html))
    blob = "\n".join(candidates)

    raw = set(u.strip() for u in JOB_URL_REGEX.findall(blob))

    resolved = set()
    for u in raw:
        # Unwrap LinkedIn safety/go?url=...
        if "linkedin.com/safety/go" in u:
            try:
                parsed = urlparse.urlparse(u)
                qs = urlparse.parse_qs(parsed.query)
                target = qs.get("url", [None])[0]
                if target:
                    target = urlparse.unquote(target)
                    resolved.add(target.strip())
                    continue
            except Exception:
                pass
        resolved.add(u)

    cleaned = set()
    for u in resolved:
        u2 = u.split("?")[0].split("#")[0].rstrip(">/).,'\"’").rstrip("/")
        if (
            "linkedin.com/jobs" in u2
            or "lnkd.in/" in u2
            or "jobindex.dk/vis-job" in u2
        ):
            cleaned.add(u2)
    return sorted(cleaned)

# =========================
# Main
# =========================
def main():
    ensure_excel(EXCEL_PATH)
    existing_urls = read_existing_urls(EXCEL_PATH)
    state = load_state(STATE_PATH)
    processed_ids = set(state.get("processed_ids", []))

    service = get_gmail_service()
    results = service.users().messages().list(userId="me", q=GMAIL_QUERY, maxResults=100).execute()
    messages = results.get("messages", [])

    next_token = results.get("nextPageToken")
    while next_token:
        more = service.users().messages().list(userId="me", q=GMAIL_QUERY, pageToken=next_token, maxResults=100).execute()
        messages.extend(more.get("messages", []))
        next_token = more.get("nextPageToken")

    new_rows, newly_processed = [], []

    for m in messages:
        msg_id = m["id"]
        if msg_id in processed_ids:
            continue

        msg = get_message(service, msg_id)
        payload = msg.get("payload", {})
        headers = payload.get("headers", [])

        subject = extract_headers(headers, "Subject") or ""
        from_ = extract_headers(headers, "From") or ""
        date_str = extract_headers(headers, "Date") or ""

        try:
            parsed = dt.datetime.strptime(date_str, "%a, %d %b %Y %H:%M:%S %z")
            display_date = parsed.astimezone().strftime("%Y-%m-%d %H:%M")
        except Exception:
            display_date = date_str

        plain, html = flattened_bodies(payload)

        # (3) use new job extractor
        urls = extract_job_urls(subject, plain, html)

        appended_any = False
        if urls:
            permalink = f"https://mail.google.com/mail/u/0/#inbox/{msg_id}"
            for u in urls:
                if u not in existing_urls:
                    # Default columns (no Source)
                    new_rows.append([display_date, from_, subject, u, permalink, "", ""])

                    # (4) Optional: include "Source" (domain). If you enable this,
                    # update HEADERS/widths/Status column range above accordingly.
                    # src = _urlp(u).netloc
                    # new_rows.append([display_date, from_, subject, u, permalink, src, "", ""])

                    existing_urls.add(u)
                    appended_any = True

        if appended_any:
            newly_processed.append(msg_id)

    if new_rows:
        append_rows(EXCEL_PATH, new_rows)

    processed_ids.update(newly_processed)
    save_state(STATE_PATH, {"processed_ids": sorted(processed_ids)})

    print(f"Done. Added {len(new_rows)} new rows to {EXCEL_PATH}.")

if __name__ == "__main__":
    main()
