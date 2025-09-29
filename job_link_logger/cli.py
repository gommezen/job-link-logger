import os
import re
import json
import base64
import datetime as dt
from typing import List, Tuple, Optional
import urllib.parse as urlparse
import argparse
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import html2text

from pathlib import Path
from job_link_logger.config import (
    EXCEL_PATH as EXCEL_PATH_DEFAULT,
    STATE_PATH as STATE_PATH_DEFAULT,
    GMAIL_QUERY as GMAIL_QUERY_DEFAULT,
    CREDENTIALS_PATH as CREDENTIALS_DEFAULT,
    TOKEN_PATH as TOKEN_DEFAULT,
    APP_DIR,
)


# Gmail API scope
SCOPES = ["https://www.googleapis.com/auth/gmail.readonly"]

# =========================
# Config
# =========================

# Unified URL matcher: LinkedIn jobs, lnkd.in, Jobindex
JOB_URL_REGEX = re.compile(
    r"(?:"  # non-capturing group
    r'https?://(?:www\.)?linkedin\.com[^\s"\'<)]+'
    r"|https?://lnkd\.in/[A-Za-z0-9_-]+"
    r'|https?://(?:www\.)?jobindex\.dk/vis-job/[^\s"\'<)]+'
    r")",
    re.IGNORECASE,
)


# =========================
# Gmail auth & client
# =========================
def get_gmail_service(credentials_path: str, token_path: str):
    """Build a Gmail API client."""
    candidates = [
        Path(credentials_path).expanduser(),
        Path.cwd() / "credentials.json",
        Path(__file__).resolve().parent / "credentials.json",
        Path(APP_DIR) / "credentials.json",
    ]
    cred_file = next((p for p in candidates if p.is_file()), None)
    if not cred_file:
        raise FileNotFoundError(
            "credentials.json not found.\n"
            "Place it in one of:\n"
            f"  - {credentials_path}\n"
            f"  - {Path.cwd() / 'credentials.json'}\n"
            f"  - {Path(__file__).resolve().parent / 'credentials.json'}\n"
            f"  - {Path(APP_DIR) / 'credentials.json'}\n"
            "Or pass --credentials <path> / set CREDENTIALS_PATH."
        )

    token_file = Path(token_path).expanduser()
    token_file.parent.mkdir(parents=True, exist_ok=True)

    creds = None
    if token_file.exists():
        creds = Credentials.from_authorized_user_file(str(token_file), SCOPES)

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            from google.auth.transport.requests import Request

            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(str(cred_file), SCOPES)
            creds = flow.run_local_server(port=0)
        with open(token_file, "w", encoding="utf-8") as f:
            f.write(creds.to_json())

    return build("gmail", "v1", credentials=creds)


# =========================
# Excel helpers
# =========================
HEADERS = [
    "Date",
    "From",
    "Subject",
    "Job URL",
    "Gmail Permalink",
    "Status",
    "Notes",
]


def ensure_excel(path: str) -> None:
    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Links"
        ws.append(HEADERS)
        widths = [22, 38, 50, 85, 60, 14, 40]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        data = wb.create_sheet("Data")
        statuses = [
            "",
            "To Review",
            "Applied",
            "Interview",
            "Offer",
            "Rejected",
            "On Hold",
        ]
        data["A1"] = "Statuses"
        for i, s in enumerate(statuses, start=2):
            data[f"A{i}"] = s

        dv = DataValidation(
            type="list",
            formula1="=Data!$A$2:$A$8",
            allow_blank=True,
        )
        ws.add_data_validation(dv)
        dv.add("F2:F10000")

        wb.save(path)


def read_existing_urls(path: str) -> set:
    wb = load_workbook(path)
    ws = wb["Links"]
    urls = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[3]:
            urls.add(str(row[3]).strip())
    return urls


def append_rows(path: str, rows: List[List[str]]) -> None:
    wb = load_workbook(path)
    ws = wb["Links"]
    for r in rows:
        ws.append(r)
    wb.save(path)


# =========================
# State
# =========================
def load_state(path: str) -> dict:
    if os.path.exists(path):
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    return {"processed_ids": []}


def save_state(path: str, state: dict) -> None:
    with open(path, "w", encoding="utf-8") as f:
        json.dump(state, f, ensure_ascii=False, indent=2)


# =========================
# Gmail parsing
# =========================
def get_message(service, msg_id: str) -> dict:
    return service.users().messages().get(userId="me", id=msg_id, format="full").execute()


def decode_b64(data: str) -> str:
    return base64.urlsafe_b64decode(data.encode("utf-8")).decode("utf-8", errors="ignore")


def extract_headers(payload_headers: List[dict], name: str) -> Optional[str]:
    for h in payload_headers:
        if h.get("name", "").lower() == name.lower():
            return h.get("value")
    return None


def flattened_bodies(payload: dict) -> Tuple[str, str]:
    plain_parts, html_parts = [], []

    def walk(part):
        mime = part.get("mimeType", "")
        body = part.get("body", {})
        data = body.get("data")
        if data:
            text = decode_b64(data)
            if mime == "text/plain":
                plain_parts.append(text)
            elif mime == "text/html":
                html_parts.append(text)
        if "parts" in part:
            for p in part["parts"]:
                walk(p)

    walk(payload)
    return ("\n".join(plain_parts), "\n".join(html_parts))


def extract_job_urls(subject: str, plain: str, html: str) -> List[str]:
    candidates = [subject or "", plain or ""]
    if html:
        candidates.append(html)
        candidates.append(html2text.html2text(html))
    blob = "\n".join(candidates)

    raw = set(u.strip() for u in JOB_URL_REGEX.findall(blob))

    resolved = set()
    for u in raw:
        if "linkedin.com/safety/go" in u:
            try:
                parsed = urlparse.urlparse(u)
                qs = urlparse.parse_qs(parsed.query)
                target = qs.get("url", [None])[0]
                if target:
                    target = urlparse.unquote(target)
                    resolved.add(target.strip())
                    continue
            except Exception:
                pass
        resolved.add(u)

    cleaned = set()
    for u in resolved:
        u2 = u.split("?")[0].split("#")[0].rstrip(">/).,'\"’").rstrip("/")
        if "linkedin.com/jobs" in u2 or "lnkd.in/" in u2 or "jobindex.dk/vis-job" in u2:
            cleaned.add(u2)
    return sorted(cleaned)


# =========================
# Core
# =========================
def main(
    excel_path: str = EXCEL_PATH_DEFAULT,
    state_path: str = STATE_PATH_DEFAULT,
    gmail_query: str = GMAIL_QUERY_DEFAULT,
    reset: bool = False,
    credentials_path: str = str(CREDENTIALS_DEFAULT),
    token_path: str = str(TOKEN_DEFAULT),
):
    service = get_gmail_service(
        credentials_path=credentials_path,
        token_path=token_path,
    )

    if reset:
        try:
            if os.path.exists(excel_path):
                os.remove(excel_path)
            if os.path.exists(state_path):
                os.remove(state_path)
            print("Reset: removed Excel and state files.")
        except Exception as e:
            print(f"Reset warning: {e}")

    ensure_excel(excel_path)
    existing_urls = read_existing_urls(excel_path)
    state = load_state(state_path)
    processed_ids = set(state.get("processed_ids", []))

    results = service.users().messages().list(userId="me", q=gmail_query, maxResults=100).execute()
    messages = results.get("messages", [])

    next_token = results.get("nextPageToken")
    while next_token:
        more = (
            service.users()
            .messages()
            .list(
                userId="me",
                q=gmail_query,
                pageToken=next_token,
                maxResults=100,
            )
            .execute()
        )
        messages.extend(more.get("messages", []))
        next_token = more.get("nextPageToken")

    new_rows, newly_processed = [], []

    for m in messages:
        msg_id = m["id"]
        if msg_id in processed_ids:
            continue

        msg = get_message(service, msg_id)
        payload = msg.get("payload", {})
        headers = payload.get("headers", [])

        subject = extract_headers(headers, "Subject") or ""
        from_ = extract_headers(headers, "From") or ""
        date_str = extract_headers(headers, "Date") or ""

        try:
            parsed = dt.datetime.strptime(
                date_str,
                "%a, %d %b %Y %H:%M:%S %z",
            )
            display_date = parsed.astimezone().strftime("%Y-%m-%d %H:%M")
        except Exception:
            display_date = date_str

        plain, html = flattened_bodies(payload)
        urls = extract_job_urls(subject, plain, html)

        appended_any = False
        if urls:
            permalink = f"https://mail.google.com/mail/u/0/#inbox/{msg_id}"
            for u in urls:
                if u not in existing_urls:
                    new_rows.append([display_date, from_, subject, u, permalink, "", ""])

                    existing_urls.add(u)
                    appended_any = True

        if appended_any:
            newly_processed.append(msg_id)

    if new_rows:
        append_rows(excel_path, new_rows)
        print(f"✅ Added {len(new_rows)} new jobs to {excel_path}.")
    else:
        print("✅ No new jobs found.")

    processed_ids.update(newly_processed)
    save_state(state_path, {"processed_ids": sorted(processed_ids)})


# ===========================
# Doctor
# =========================
def doctor(
    credentials_path: str,
    token_path: str,
    excel_path: str,
    state_path: str,
):
    print("🔍 Job Link Logger Doctor\n")

    cred = Path(credentials_path).expanduser()
    print(f"{'✅' if cred.exists() else '❌'} " f"credentials.json at {cred}")

    token = Path(token_path).expanduser()
    print(f"{'✅' if token.exists() else '⚠️'} " f"token.json at {token}")

    excel = Path(excel_path)
    if excel.exists():
        try:
            urls = read_existing_urls(excel)
            print(f"✅ Excel file at {excel} " f"({len(urls)} job links)")
        except Exception as e:
            print(f"❌ Excel file exists but could not be read: {e}")
    else:
        print("⚠️ Excel file not found " "(will be created on next run)")

    state = Path(state_path)
    if state.exists():
        try:
            with open(state, "r", encoding="utf-8") as f:
                data = json.load(f)
            count = len(data.get("processed_ids", []))
            print(f"✅ State file at {state} " f"({count} Gmail messages processed)")
        except Exception as e:
            print(f"❌ State file exists but could not be read: {e}")
    else:
        print("⚠️ State file not found " "(will be created on first successful run)")


# =========================
# CLI entry
# =========================
def run():
    parser = argparse.ArgumentParser(
        prog="job-link-logger",
        description=("Log LinkedIn / Jobindex job links " "from Gmail into Excel."),
    )

    subparsers = parser.add_subparsers(dest="command")

    run_parser = subparsers.add_parser("run", help="Run the job link logger")
    run_parser.add_argument("--excel", default=EXCEL_PATH_DEFAULT)
    run_parser.add_argument("--state", default=STATE_PATH_DEFAULT)
    run_parser.add_argument("--query", default=GMAIL_QUERY_DEFAULT)
    run_parser.add_argument("--credentials", default=str(CREDENTIALS_DEFAULT))
    run_parser.add_argument("--token", default=str(TOKEN_DEFAULT))
    run_parser.add_argument("--reset", action="store_true")

    doc_parser = subparsers.add_parser("doctor", help="Check config and files")
    doc_parser.add_argument("--credentials", default=str(CREDENTIALS_DEFAULT))
    doc_parser.add_argument("--token", default=str(TOKEN_DEFAULT))
    doc_parser.add_argument("--excel", default=EXCEL_PATH_DEFAULT)
    doc_parser.add_argument("--state", default=STATE_PATH_DEFAULT)

    args = parser.parse_args()

    if args.command == "doctor":
        doctor(args.credentials, args.token, args.excel, args.state)
    else:
        main(
            excel_path=getattr(args, "excel", EXCEL_PATH_DEFAULT),
            state_path=getattr(args, "state", STATE_PATH_DEFAULT),
            gmail_query=getattr(args, "query", GMAIL_QUERY_DEFAULT),
            credentials_path=getattr(args, "credentials", str(CREDENTIALS_DEFAULT)),
            token_path=getattr(args, "token", str(TOKEN_DEFAULT)),
            reset=getattr(args, "reset", False),
        )
